import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
import xlwings as xw

AURA_FILE = "Aura.xquant"

# --- Sheets Config ---
core_hub = ["Simulation_Scenarios", "Visualization_Config", "Deployment",
            "Ethics_Notes", "Collaboration_Log"]

stem_sheets = ["Advanced_Mathematics", "Physics_Experiments", "Reasoning_Problems",
               "Genomics_Deep", "Healthcare_Analytics", "Environment_Scenarios",
               "AI_Results_Log"]

# --- Step 1: Create workbook (xlsx first) ---
wb = openpyxl.Workbook()
home = wb.active
home.title = "Home"

# Title Block
home.merge_cells("A1:E2")
cell = home["A1"]
cell.value = "🌌 Aura Hub – Research Ecosystem"
cell.font = Font(size=16, bold=True, color="FFFFFF")
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

# Navigation placeholders
row = 5
for s in core_hub:
    home[f"A{row}"].value = s
    home[f"A{row}"].hyperlink = f"#{s}!A1"
    home[f"A{row}"].style = "Hyperlink"
    row += 1

row = 5
for s in stem_sheets:
    home[f"C{row}"].value = s
    home[f"C{row}"].hyperlink = f"#{s}!A1"
    home[f"C{row}"].style = "Hyperlink"
    row += 1

quick_actions = ["▶️ Run Simulation", "📈 View Charts", "🚀 Deploy Configs",
                 "📒 Open Ethics Notes", "🤝 Collaboration"]
row = 5
for q in quick_actions:
    home[f"E{row}"].value = q
    row += 1

# Info Panel
home.merge_cells("A15:E18")
cell = home["A15"]
cell.value = (
    "ℹ️ Aura Hub centralizes:\n"
    "• STEM Research Sheets (Math, Physics, Genomics, Healthcare, Environment)\n"
    "• AI & Quantum Simulation Results\n"
    "• Ethics, Economics & Deployment Tracking\n\n"
    "🔗 Version: 1.0 | Last Updated: 2025-09-19"
)
cell.alignment = Alignment(wrap_text=True, vertical="top")

# Adjust cols
for col in range(1, 6):
    home.column_dimensions[get_column_letter(col)].width = 25

# Create other sheets
for s in core_hub + stem_sheets:
    ws = wb.create_sheet(title=s)
    ws["A1"].value = f"Sheet: {s}"

# Save as xlsx first
tmp_file = "Aura_tmp.xlsx"
wb.save(tmp_file)

# --- Step 2: Open with xlwings and inject VBA ---
app = xw.App(visible=False)
book = xw.Book(tmp_file)

vba_code = '''
Sub RunSimulation()
    MsgBox "🔬 Running Simulation Scenarios...", vbInformation, "Aura Hub"
    Sheets("Simulation_Scenarios").Activate
End Sub

Sub ViewCharts()
    MsgBox "📈 Generating Charts from Visualization_Config...", vbInformation, "Aura Hub"
    Sheets("Visualization_Config").Activate
End Sub

Sub DeployConfigs()
    MsgBox "🚀 Loading Deployment Settings...", vbInformation, "Aura Hub"
    Sheets("Deployment").Activate
End Sub

Sub OpenEthics()
    Sheets("Ethics_Notes").Activate
End Sub

Sub OpenCollab()
    Sheets("Collaboration_Log").Activate
End Sub

Sub ResetHub()
    MsgBox "♻️ Aura Hub has been reset.", vbExclamation, "Aura Hub"
    Sheets("Home").Activate
End Sub

Sub HubStatus()
    MsgBox "✅ Aura Hub is running normally.", vbInformation, "Aura Hub"
End Sub

Sub CreateHomeButtons()
    Dim btn As Shape
    Dim i As Integer
    Dim actions As Variant, macros As Variant
    
    actions = Array("▶️ Run Simulation", "📈 View Charts", "🚀 Deploy Configs", "📒 Open Ethics Notes", "🤝 Collaboration")
    macros = Array("RunSimulation", "ViewCharts", "DeployConfigs", "OpenEthics", "OpenCollab")
    
    For i = LBound(actions) To UBound(actions)
        Set btn = Sheets("Home").Shapes.AddShape(msoShapeRoundedRectangle, 400, 80 + i * 40, 200, 30)
        btn.TextFrame.Characters.Text = actions(i)
        btn.OnAction = macros(i)
        btn.Fill.ForeColor.RGB = RGB(79, 129, 189)
        btn.TextFrame.HorizontalAlignment = xlHAlignCenter
        btn.TextFrame.VerticalAlignment = xlVAlignCenter
        btn.TextFrame.Characters.Font.Color = vbWhite
        btn.TextFrame.Characters.Font.Bold = True
    Next i
End Sub
'''

book.api.VBProject.VBComponents.Add(1).CodeModule.AddFromString(vba_code)

# Save as macro-enabled
book.save(AURA_FILE)
book.close()
app.quit()

print(f"✅ Aura Hub with VBA saved as {AURA_FILE}")
