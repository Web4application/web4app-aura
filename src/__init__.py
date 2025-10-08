from openpyxl import load_workbook
from openpyxl.styles import Font

# Load the existing workbook
file_path = "/mnt/data/Aura_Hub.xlsx"
wb = load_workbook(file_path)

# Ensure Home sheet exists
if "Home" not in wb.sheetnames:
    home = wb.create_sheet("Home")
else:
    home = wb["Home"]

# Clear existing content in Home
for row in home["A1:A30"]:
    for cell in row:
        cell.value = None

# Title
home["A1"] = "🌌 Aura Hub – Research Ecosystem"
home["A1"].font = Font(size=16, bold=True)

# Define hyperlinks with emojis
links = [
    ("🌐 GitHub Repository", "https://web4application.github.io/Aura/"),
    ("📄 Documentation", "https://github.com/web4application/Aura"),
    ("📝 Experiment Logs", "#Logs!A1"),
    ("📊 Data Sheets", "#Data!A1"),
    ("🤝 Collaboration Log", "#Collaboration_Log!A1"),
    ("📈 Visualization Config", "#Visualization_Config!A1"),
    ("🚀 Deployment Settings", "#Deployment!A1"),
]

# Insert links into Home sheet
row = 3
for name, url in links:
    cell = home[f"A{row}"]
    cell.value = name
    cell.hyperlink = url
    cell.font = Font(size=12, color="0000FF", underline="single")
    row += 2

# Save new file
updated_path = "/mnt/data/Aura_Hub_Hyperlinks.xlsx"
wb.save(updated_path)
updated_path
