import openpyxl

class SERAIEngine:
    def __init__(self, workbook_path):
        self.workbook_path = workbook_path
        self.wb = openpyxl.load_workbook(workbook_path)
    
    def list_sheets(self):
        return self.wb.sheetnames

    def read_sheet(self, sheet_name):
        if sheet_name not in self.wb.sheetnames:
            return None
        sheet = self.wb[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        return data

    def summarize_tp_stages(self):
        if "Teleportation_Simulation" not in self.wb.sheetnames:
            return "No teleportation data found."
        sheet = self.wb["Teleportation_Simulation"]
        summary = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            stage, description, status = row[0], row[1], row[2]
            summary[stage] = {"description": description, "status": status}
        return summary


if __name__ == "__main__":
    engine = SERAIEngine("../extensions/Aura.xlsx")
    print("Sheets:", engine.list_sheets())
    tp_summary = engine.summarize_tp_stages()
    for stage, info in tp_summary.items():
        print(f"{stage}: {info}")
