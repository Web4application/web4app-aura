from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Load existing workbook
file_path = "/mnt/data/Aura_Hub.xlsx"
wb = load_workbook(file_path)

# Select Home sheet
home = wb["Home"]

# Clear existing content before adding new structured layout
for row in home["A1:D20"]:
    for cell in row:
        cell.value = None

# Title
home["A1"] = "🌌 Aura Hub - Research Ecosystem"
home["A1"].font = Font(size=16, bold=True)

# Section for navigation
links = {
    "GitHub Repository": "https://web4application.github.io/Aura/",
    "Documentation": "https://github.com/web4application/Aura",
    "Experiment Logs": "#Logs!A1",
    "Data Sheets": "#Data!A1"
}

row = 3
for name, url in links.items():
    cell = home[f"A{row}"]
    cell.value = name
    cell.font = Font(size=12, color="0000FF", underline="single")
    home[f"A{row}"].hyperlink = url
    row += 2

# Adjust column width
for col in range(1, 3):
    home.column_dimensions[get_column_letter(col)].width = 40

# Save updated workbook
updated_path = "/mnt/data/Aura_Hub_with_links.xlsx"
wb.save(updated_path)
updated_path