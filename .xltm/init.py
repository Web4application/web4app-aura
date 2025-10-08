import openpyxl

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Simulation_Problems"

# SA placeholders
ws['A1'] = "SA chosen path"
ws['B1'] = ""
ws['A2'] = "SA objective"
ws['B2'] = ""

# QAOA placeholders
ws['A4'] = "QAOA chosen path"
ws['B4'] = ""
ws['A5'] = "QAOA objective"
ws['B5'] = ""

# Save as .xlsx first
wb.save("Aura_telecom_simulation.xlsx")

Sub RunAuraTelecom()
    RunPython ("import telecom_aura; telecom_aura.update_aura_xlsl()")
End Sub
